import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

def save_excel_as_csv_custom(xlsx_path):
    downloads_dir = Path.home() / "Downloads"

    # 打开 Excel 获取 A2 日期
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]
    date_cell = ws["A2"].value
    if not date_cell:
        raise ValueError("第一个工作表的 A2 单元格为空")

    # 日期格式：补零（2025-09-01）
    if hasattr(date_cell, "year") and hasattr(date_cell, "month") and hasattr(date_cell, "day"):
        date_str = f"{date_cell.year:04d}-{date_cell.month:02d}-{date_cell.day:02d}"
    else:
        date_str = str(date_cell)

    csv_name = f"FBA日记 {date_str}.csv"
    csv_path = downloads_dir / csv_name

    # pandas 读取第一个工作表
    df = pd.read_excel(xlsx_path, sheet_name=ws.title)

    # 把 0 转为空字符串，空单元格保持空
    df = df.replace(0, "")

    # P列（第16列）保留 1 位小数
    if df.shape[1] >= 16:
        col_p = df.columns[15]  # Python 0-index
        df[col_p] = df[col_p].apply(lambda x: f"{x:.1f}" if pd.notnull(x) and x != "" else "")

    # 保存 CSV
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    print(f"✅ 已保存 CSV: {csv_path}")

# 调用示例
save_excel_as_csv_custom(r"C:\ACT\公用核心\2.8_FBA日记.xlsx")
