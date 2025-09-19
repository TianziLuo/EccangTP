from openpyxl import load_workbook

def delete_columns_in_sheet(filepath, sheetname="读取原始记录"):
    # 打开工作簿
    wb = load_workbook(filepath)
    ws = wb[sheetname]

    # 找到最后一行
    max_row = ws.max_row

    # 删除 F 到 CW 列，第 2 行到最后一行
    for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=6, max_col=101):
        for cell in row:
            cell.value = None

    # 保存
    wb.save(filepath)
    print(f"✅ 已清空 {sheetname} 的 F-CW 列（第2行到第{max_row}行）")

if __name__ == "__main__":
    path = r"C:\ACT\公用核心\2.8_FBA日记.xlsx"
    delete_columns_in_sheet(path)
