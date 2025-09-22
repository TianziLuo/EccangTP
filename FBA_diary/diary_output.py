import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from config_paths import get_diary_paths


paths = get_diary_paths()


def save_excel_as_csv_custom(xlsx_path):
    # Target directory to save the CSV
    target_dir = paths["diary_folder"]

    # Load the workbook and get the first worksheet
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]

    # Read the date from cell A2
    date_cell = ws["A2"].value
    if not date_cell:
        raise ValueError("Cell A2 in the first worksheet is empty")

    # Format date as YYYY-MM-DD
    if hasattr(date_cell, "year") and hasattr(date_cell, "month") and hasattr(date_cell, "day"):
        date_str = f"{date_cell.year:04d}-{date_cell.month:02d}-{date_cell.day:02d}"
    else:
        # If not a datetime object, replace slashes with dashes
        date_str = str(date_cell).replace("/", "-").replace("\\", "-")

    # Build the CSV file name
    csv_name = f"FBA日记 {date_str}.csv"
    csv_path = target_dir / csv_name

    # Read the first worksheet
    df = pd.read_excel(xlsx_path, sheet_name=ws.title)

    # Replace numeric zeros with empty strings
    df = df.applymap(lambda x: "" if (isinstance(x, (int, float)) and x == 0) else x)

    # Keep only rows where column B or C has content
    if df.shape[1] >= 3:
        df = df[df.iloc[:, 1].notna() | df.iloc[:, 2].notna()]

    # Column P (16th column) → keep 1 decimal place if exists
    if df.shape[1] >= 16:
        col_p = df.columns[15]
        df[col_p] = df[col_p].apply(lambda x: f"{x:.1f}" if pd.notnull(x) and x != "" else "")

    # Save DataFrame as CSV
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    print(f"✅ CSV saved: {csv_path}")
